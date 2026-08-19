"""
compare_workflow_template_vs_ddm.py

Purpose
-------
For every APPROVED document in the local Document Register table:
    1. Get its Document Number and PLIP ID.
    2. Look up its Workflow instances in the local Workflow table (using
       Document Number), and resolve which single Workflow Template
       (Aconex) was actually used for that document's workflow.
    3. Look up that PLIP ID in the DDM Excel export and read the Template
       column defined there.
    4. Compare the two template names and report Match / Mismatch /
       Not Found / Ambiguous.

Output
------
An Excel report (+ console summary) listing, per document:
    Document Number | PLIP ID | Aconex Template | DDM Template | Result | Notes

Everything you are likely to need to change (DB connection, table names,
column names, DDM file/sheet, output path) lives in the CONFIG block
below. Nothing else in the script should need editing for routine use.

Requirements
------------
    pip install pyodbc pandas openpyxl
    (pyodbc requires the appropriate ODBC driver for your DB installed,
     e.g. "ODBC Driver 17 for SQL Server")
"""

import re
import sys
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import pyodbc
except ImportError:
    pyodbc = None


# ============================================================================
# CONFIG — EDIT THIS SECTION ONLY
# ============================================================================


class Config:
    # ---- Database connection -------------------------------------------------
    DB_DRIVER = "{ODBC Driver 18 for SQL Server}"
    DB_SERVER = "es-mssql-01"  # e.g. "SERVERNAME\\SQLEXPRESS" or an IP
    DB_NAME = "ACONEX Reporting Data"
    DB_TRUSTED_CONNECTION = True  # True = use Windows auth, ignore user/pass below
    DB_USERNAME = ""
    DB_PASSWORD = ""

    # ---- Project Filter ------------------------------------------------------
    PROJECT_ID = 1342183550  # change per project

    # ---- Document Register table ----------------------------------------------
    DOC_REGISTER_TABLE = "DocumentRegisterHistory"
    DOC_REGISTER_COL_DOC_NUMBER = "DocNo"
    DOC_REGISTER_COL_REVISION = "Revision"
    DOC_REGISTER_COL_STATUS = "ReviewStatus"
    DOC_REGISTER_COL_PLIP_ID = "VDRCode"
    APPROVED_STATUS_VALUE = "Approved"  # exact string that means "Approved" in your DB
    DOC_REGISTER_COL_PROJECT_ID = "ProjectId"

    # ---- Workflow table ---------------------------------------------------------
    WORKFLOW_TABLE = "Workflows"
    WORKFLOW_COL_DOC_NUMBER = "DocumentNumber"
    WORKFLOW_COL_REVISION = "DocumentRevision"
    WORKFLOW_COL_WORKFLOW_NUMBER = "WorkflowNumber"  # groups step-instances together
    WORKFLOW_COL_TEMPLATE_NAME = "WorkflowTemplate"
    WORKFLOW_COL_PROJECT_ID = "ProjectId"
    WORKFLOW_COL_STATUS = "WorkflowStatus"

    # ---- DDM Excel file -----------------------------------------------------------
    DDM_FILE_PATH = ""
    DDM_SHEET_NAME = 0  # sheet name or index; 0 = first sheet
    DDM_COL_PLIP_ID = "PLIP ID"
    DDM_COL_TEMPLATE_NAME = "Template"
    DDM_COL_WORKFLOW_RULE = "Column6"
    DDM_HEADER_ROW = 1  # 0-indexed row number where headers live

    # ---- Refresh Tracking ------------------------------------------------------
    ORG_ID = 1342190259
    REPORT_CATEGORY = "DDR_Check"

    REFRESH_TIME_TABLE = "dbo.RefreshTime"

    REFRESH_COL_LAST_REFRESHED = "Last Refreshed"
    REFRESH_COL_PROJECT_ID = "ProjectId"
    REFRESH_COL_REPORT_CATEGORY = "ReportCategory"
    REFRESH_COL_ORG_ID = "OrgId"

    # ---- Output ---------------------------------------------------------------------
    OUTPUT_DIR = r"\\pd-file-srv-01\Docs\AIS\DMS\Aconex\Power BI\SQL-PowerBi Report\BUDOUR (5376200)\Data for Reports"  # folder to write the report into
    OUTPUT_FILENAME_PREFIX = "workflow_template_vs_ddm_report"

    # ---- Behaviour ------------------------------------------------------------------
    # If a document's workflow instances reference more than one distinct
    # template name, should we treat that as an error (Ambiguous) or just
    # take the first one found?
    TREAT_MULTIPLE_TEMPLATES_AS_AMBIGUOUS = False

    # In Aconex, template names are built as "<DDM Template Name> (<Discipline>)"
    # e.g. Aconex: "8804-Project Sub system/existing sytem Interface Tie - In Pack
    #               document (Telecommunication)"
    #      DDM:    "8804-Project Sub system/existing sytem Interface Tie - In Pack
    #               document"
    # When True, a trailing " (....)" tag is stripped from the Aconex template
    # name before comparing it against the DDM template name.
    STRIP_TRAILING_DISCIPLINE_TAG_FROM_ACONEX = True

    LOG_LEVEL = logging.INFO


# ============================================================================
# END OF CONFIG
# ============================================================================


logging.basicConfig(
    level=Config.LOG_LEVEL,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# Matches a single trailing "(...)" group at the very end of the string,
# with optional surrounding whitespace, e.g.:
#   "...Pack document (Telecommunication)" -> "...Pack document"
#   "...Package Data Sheet (Pressure Vessels)" -> "...Package Data Sheet"
TRAILING_BRACKET_RE = re.compile(r"\s*\([^()]*\)\s*$")


def strip_trailing_discipline_tag(template_name: str) -> str:
    """Remove a trailing ' (Discipline)' tag from an Aconex template name."""
    if not template_name:
        return template_name
    return TRAILING_BRACKET_RE.sub("", template_name).strip()


@dataclass
class ComparisonResult:
    document_number: str
    revision: str
    plip_id: str
    ddm_template: str
    workflow_rule: str
    workflows_found: str
    has_required_workflow: str
    has_pem_approval: str
    result: str
    notes: str = ""


def build_connection():
    """Open a pyodbc connection using the CONFIG settings above."""
    if pyodbc is None:
        log.error("pyodbc is not installed. Run: pip install pyodbc")
        sys.exit(1)

    if Config.DB_TRUSTED_CONNECTION:
        conn_str = (
            f"DRIVER={Config.DB_DRIVER};"
            f"SERVER={Config.DB_SERVER};"
            f"DATABASE={Config.DB_NAME};"
            f"Trusted_Connection=yes;"
            "Encrypt=no;"
        )
    else:
        conn_str = (
            f"DRIVER={Config.DB_DRIVER};"
            f"SERVER={Config.DB_SERVER};"
            f"DATABASE={Config.DB_NAME};"
            f"UID={Config.DB_USERNAME};"
            f"PWD={Config.DB_PASSWORD};"
        )

    log.info("Connecting to database %s on %s ...", Config.DB_NAME, Config.DB_SERVER)
    return pyodbc.connect(conn_str)


def get_approved_documents(conn) -> pd.DataFrame:
    """Return DataFrame of [doc_number, plip_id] for all Approved documents."""
    query = f"""
        SELECT
            [{Config.DOC_REGISTER_COL_DOC_NUMBER}] AS doc_number,
            [{Config.DOC_REGISTER_COL_REVISION}] AS revision,
            [{Config.DOC_REGISTER_COL_PLIP_ID}] AS plip_id
        FROM {Config.DOC_REGISTER_TABLE}
        WHERE [{Config.DOC_REGISTER_COL_STATUS}] = ?
        AND [{Config.DOC_REGISTER_COL_PROJECT_ID}] = ?
    """
    log.info("Fetching Approved documents from %s ...", Config.DOC_REGISTER_TABLE)
    df = pd.read_sql(
        query, conn, params=[Config.APPROVED_STATUS_VALUE, Config.PROJECT_ID]
    )
    df["doc_number"] = df["doc_number"].astype(str).str.strip()
    df["revision"] = df["revision"].astype(str).str.strip()
    df["plip_id"] = df["plip_id"].astype(str).str.strip()
    log.info("Found %d Approved document(s).", len(df))
    return df


def get_workflow_templates_for_document(conn, doc_number: str, revision: str):

    query = f"""
        SELECT DISTINCT
            [{Config.WORKFLOW_COL_TEMPLATE_NAME}] AS template_name
        FROM {Config.WORKFLOW_TABLE}
        WHERE [{Config.WORKFLOW_COL_DOC_NUMBER}] = ?
        AND [{Config.WORKFLOW_COL_REVISION}] = ?
        AND [{Config.WORKFLOW_COL_PROJECT_ID}] = ?
        AND ISNULL([{Config.WORKFLOW_COL_STATUS}], '') <> 'Terminated'
    """

    df = pd.read_sql(query, conn, params=[doc_number, revision, Config.PROJECT_ID])

    if df.empty:
        return []

    templates = df["template_name"].astype(str).str.strip().tolist()

    templates = [t for t in templates if t and t.lower() != "none"]

    return list(set(templates))


def load_ddm() -> pd.DataFrame:
    """Load the DDM excel sheet and return a lookup keyed by PLIP ID."""
    log.info("Loading DDM file: %s", Config.DDM_FILE_PATH)
    ddm_path = Path(Config.DDM_FILE_PATH)
    if not ddm_path.exists():
        log.error("DDM file not found at %s", ddm_path)
        sys.exit(1)

    df = pd.read_excel(
        ddm_path,
        sheet_name=Config.DDM_SHEET_NAME,
        header=Config.DDM_HEADER_ROW,
    )

    missing = [
        c
        for c in (
            Config.DDM_COL_PLIP_ID,
            Config.DDM_COL_TEMPLATE_NAME,
            Config.DDM_COL_WORKFLOW_RULE,
        )
        if c not in df.columns
    ]

    if missing:
        log.error(
            "DDM file is missing expected column(s): %s. Columns found: %s",
            missing,
            list(df.columns),
        )
        sys.exit(1)

    df[Config.DDM_COL_PLIP_ID] = df[Config.DDM_COL_PLIP_ID].astype(str).str.strip()
    df[Config.DDM_COL_TEMPLATE_NAME] = (
        df[Config.DDM_COL_TEMPLATE_NAME].astype(str).str.strip()
    )

    log.info("DDM loaded: %d row(s).", len(df))
    return df.set_index(Config.DDM_COL_PLIP_ID)


def compare(conn, ddm_lookup, approved_docs):

    results = []

    for _, row in approved_docs.iterrows():

        doc_number = row["doc_number"]
        revision = row["revision"]
        plip_id = row["plip_id"]

        workflow_templates = get_workflow_templates_for_document(
            conn, doc_number, revision
        )

        if plip_id not in ddm_lookup.index:

            results.append(
                ComparisonResult(
                    document_number=doc_number,
                    revision=revision,
                    plip_id=plip_id,
                    ddm_template="",
                    workflow_rule="",
                    workflows_found=", ".join(workflow_templates),
                    has_required_workflow="N/A",
                    has_pem_approval="N/A",
                    result="DDM NOT FOUND",
                    notes="PLIP ID not found in DDM",
                )
            )

            continue

        ddm_row = ddm_lookup.loc[plip_id]

        if isinstance(ddm_row, pd.DataFrame):
            ddm_row = ddm_row.iloc[0]

        ddm_template = str(ddm_row[Config.DDM_COL_TEMPLATE_NAME]).strip()

        workflow_rule = str(ddm_row[Config.DDM_COL_WORKFLOW_RULE]).strip()

        # --------------------------------------------------
        # CANNOT BE DETERMINED
        # --------------------------------------------------

        if workflow_rule.upper() in ("EMPTY", "ONLY R"):

            results.append(
                ComparisonResult(
                    document_number=doc_number,
                    revision=revision,
                    plip_id=plip_id,
                    ddm_template=ddm_template,
                    workflow_rule=workflow_rule,
                    workflows_found=", ".join(workflow_templates),
                    has_required_workflow="N/A",
                    has_pem_approval="N/A",
                    result="Can't Be Determined",
                    notes=f"Column6 value is '{workflow_rule}'.",
                )
            )

            continue

        normalized_templates = [
            strip_trailing_discipline_tag(t).lower() for t in workflow_templates
        ]

        has_pem_approval = any(t.lower() == "pem approval" for t in workflow_templates)

        has_required_workflow = any(
            t == ddm_template.lower() for t in normalized_templates
        )

        # --------------------------------------------------
        # ONLY O
        # --------------------------------------------------

        if workflow_rule.upper() in ("EMPTY", "ONLY R", "", "NAN"):

            if has_pem_approval:

                result = "PASS"

                notes = "PEM Approval found."

            else:

                result = "FAIL"

                notes = "Expected PEM Approval workflow but none found."

        # --------------------------------------------------
        # WORKFLOW REQUIRED
        # --------------------------------------------------

        else:

            if has_required_workflow and has_pem_approval:

                result = "PASS"

                notes = ""

            elif not has_required_workflow and has_pem_approval:

                result = "FAIL"

                notes = "Document entered PEM Approval " "without required workflow."

            elif has_required_workflow and not has_pem_approval:

                result = "FAIL"

                notes = "Required workflow found but " "PEM Approval missing."

            else:

                result = "FAIL"

                notes = "Neither required workflow " "nor PEM Approval found."

        results.append(
            ComparisonResult(
                document_number=doc_number,
                revision=revision,
                plip_id=plip_id,
                ddm_template=ddm_template,
                workflow_rule=workflow_rule,
                workflows_found=", ".join(workflow_templates),
                has_required_workflow="Yes" if has_required_workflow else "No",
                has_pem_approval="Yes" if has_pem_approval else "No",
                result=result,
                notes=notes,
            )
        )

    return results


def write_report(results: list[ComparisonResult], output_dir: str) -> Path:
    df = pd.DataFrame([r.__dict__ for r in results])
    df = df.rename(
        columns={
            "document_number": "Document Number",
            "revision": "Revision",
            "plip_id": "PLIP ID",
            "ddm_template": "DDM Template",
            "workflow_rule": "Column6",
            "workflows_found": "Workflow Templates Found",
            "has_required_workflow": "Required Workflow Found",
            "has_pem_approval": "PEM Approval Found",
            "result": "Result",
            "notes": "Notes",
        }
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path(output_dir) / f"{Config.OUTPUT_FILENAME_PREFIX}_{timestamp}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:

        df.to_excel(writer, index=False, sheet_name="Comparison")

        summary = (
            df["Result"].value_counts().rename_axis("Result").reset_index(name="Count")
        )

        summary.to_excel(writer, index=False, sheet_name="Summary")

        workbook = writer.book

        comparison_ws = writer.sheets["Comparison"]
        summary_ws = writer.sheets["Summary"]

        green_fill = PatternFill(fill_type="solid", fgColor="00B050")

        red_fill = PatternFill(fill_type="solid", fgColor="C00000")

        yellow_fill = PatternFill(fill_type="solid", fgColor="FFD966")

        orange_fill = PatternFill(fill_type="solid", fgColor="ED7D31")

        for row in range(2, summary_ws.max_row + 1):

            result = summary_ws.cell(row=row, column=1).value

            if result == "PASS":
                summary_ws.cell(row=row, column=1).fill = green_fill

            elif result == "FAIL":
                summary_ws.cell(row=row, column=1).fill = red_fill

            elif result == "Can't Be Determined":
                summary_ws.cell(row=row, column=1).fill = yellow_fill

            elif result == "DDM NOT FOUND":
                summary_ws.cell(row=row, column=1).fill = orange_fill
        # ====================================
        # Header Style
        # ====================================

        header_fill = PatternFill("solid", fgColor="1F4E78")

        header_font = Font(color="FFFFFF", bold=True)

        for ws in [comparison_ws, summary_ws]:

            for cell in ws[1]:

                cell.fill = header_fill
                cell.font = header_font

        # ====================================
        # Freeze Top Row
        # ====================================

        comparison_ws.freeze_panes = "A2"

        # ====================================
        # Autofilter
        # ====================================

        table = Table(displayName="ComparisonTable", ref=comparison_ws.dimensions)

        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style

        comparison_ws.add_table(table)
        # =====================
        # Auto Width
        # ====================================

        for ws in [comparison_ws, summary_ws]:

            for column in ws.columns:

                max_length = 0

                column_letter = get_column_letter(column[0].column)

                for cell in column:

                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[column_letter].width = min(max_length + 3, 60)

        # ====================================
        # Conditional Coloring
        # ====================================

        green_fill = PatternFill(fill_type="solid", fgColor="00B050")

        red_fill = PatternFill(fill_type="solid", fgColor="FF0000")

        yellow_fill = PatternFill(fill_type="solid", fgColor="FFD966")

        orange_fill = PatternFill(fill_type="solid", fgColor="ED7D31")

        result_col = None

        for cell in comparison_ws[1]:

            if cell.value == "Result":

                result_col = cell.column
                break

        if result_col:

            for row in range(2, comparison_ws.max_row + 1):

                result = comparison_ws.cell(row=row, column=result_col).value

                fill = None

                if result == "PASS":
                    fill = green_fill

                elif result == "FAIL":
                    fill = red_fill

                elif result == "Can't Be Determined":
                    fill = yellow_fill

                elif result == "DDM NOT FOUND":
                    fill = orange_fill

                if fill:

                    comparison_ws.cell(row=row, column=result_col).fill = fill

    return out_path


def generate_report(output_dir):

    conn = build_connection()

    try:
        approved_docs = get_approved_documents(conn)

        if approved_docs.empty:
            log.warning("No Approved documents found.")
            return None

        ddm_lookup = load_ddm()

        results = compare(conn, ddm_lookup, approved_docs)

    finally:
        conn.close()

    out_path = write_report(results, output_dir)

    # Update refresh table here
    conn = build_connection()

    try:
        update_refresh_time(conn)
    finally:
        conn.close()

    return out_path


def update_refresh_time(conn):

    now = datetime.now()

    cursor = conn.cursor()

    cursor.execute(
        f"""
        UPDATE {Config.REFRESH_TIME_TABLE}
        SET [{Config.REFRESH_COL_LAST_REFRESHED}] = ?
        WHERE [{Config.REFRESH_COL_PROJECT_ID}] = ?
          AND [{Config.REFRESH_COL_REPORT_CATEGORY}] = ?
          AND [{Config.REFRESH_COL_ORG_ID}] = ?
        """,
        (
            now,
            Config.PROJECT_ID,
            Config.REPORT_CATEGORY,
            Config.ORG_ID,
        ),
    )

    if cursor.rowcount == 0:

        cursor.execute(
            f"""
            INSERT INTO {Config.REFRESH_TIME_TABLE}
            (
                [{Config.REFRESH_COL_LAST_REFRESHED}],
                [{Config.REFRESH_COL_PROJECT_ID}],
                [{Config.REFRESH_COL_REPORT_CATEGORY}],
                [{Config.REFRESH_COL_ORG_ID}]
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                now,
                Config.PROJECT_ID,
                Config.REPORT_CATEGORY,
                Config.ORG_ID,
            ),
        )

    conn.commit()

    log.info("RefreshTime updated successfully.")


def main():
    conn = build_connection()
    try:
        approved_docs = get_approved_documents(conn)
        if approved_docs.empty:
            log.warning("No Approved documents found. Nothing to compare.")
            return

        ddm_lookup = load_ddm()
        results = compare(conn, ddm_lookup, approved_docs)
    finally:
        conn.close()

    out_path = write_report(results, Config.OUTPUT_DIR)

    # Update refresh tracking table
    log.info("Starting RefreshTime update...")

    conn = build_connection()

    try:
        update_refresh_time(conn)
    except Exception:
        log.exception("RefreshTime update failed")
    finally:
        conn.close()

    log.info("Finished RefreshTime update.")

    total = len(results)
    matches = sum(1 for r in results if r.result == "MATCH")
    mismatches = sum(1 for r in results if r.result == "MISMATCH")
    not_found = sum(1 for r in results if r.result == "NOT FOUND")

    log.info("=" * 60)
    log.info("Total documents checked : %d", total)
    log.info("Matches                 : %d", matches)
    log.info("Mismatches              : %d", mismatches)
    log.info("Not found               : %d", not_found)
    log.info("Report written to       : %s", out_path)
    import os

    os.startfile(out_path)

    log.info("=" * 60)


if __name__ == "__main__":
    main()
